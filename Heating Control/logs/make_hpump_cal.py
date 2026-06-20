"""
H-pump SOC-test calibration analyser.

Reads all sct*.csv files in the same directory as this script.

sct.csv format (no header row, written by socTestLogRow() in h_controller):
  timestamp, pct, hot_pipe, htr1, htr2, pump_pct, pred, event

Processing per step:
  STEP_START event  → start of 15-min test at one heater level
  10s prime period  → skip first 5 rows (pump forced to 100%, discard)
  Stabilisation     → wait for both htr1 and htr2 to each have a range < 0.5°C
                      over 30 consecutive seconds (15 rows at 2s/row)
  Recording         → once stable, accumulate readings per hp_nom bucket;
                      at step end (STEP_END / PAUSE / OVERHEAT / next STEP_START)
                      commit the per-bucket averages as one step entry each
  Result            → data[pct][hp_nom] = [(avg_htr, avg_pump), ...] one per step

Output: hpump_cal.xlsx
  20 sheets, one per heater level (named "5%", "10%", ... "100%")
  Row 1   : headers
  Per HP temp (15-85°C), two rows:
    Row A : "htr out" | overall avg htr | 85 | 85 | step1_avg_htr | 85 | step2_avg_htr | 85 | ...
    Row B : HP temp   | overall avg pump | lin_85 | pl_85 | step1_avg_pump | step1_lin_85 | ...
  Steps are in the order they were collected (chronological across files).
"""

import os, math
from collections import defaultdict
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

LOG_DIR  = os.path.dirname(os.path.abspath(__file__))
OUT_PATH = os.path.join(LOG_DIR, "hpump_cal.xlsx")

HP_MIN       = 15    # minimum hot-pipe nominal (°C)
HP_MAX       = 85    # maximum hot-pipe nominal (°C)
PRIME_ROWS   = 5     # rows to skip after STEP_START (10 s / 2 s per row)
STABLE_ROWS  = 15    # consecutive rows required for stability (30 s)
STABLE_THR   = 0.5   # °C — each sensor's own range must be below this

# data[pct][hp_nom] = [(avg_htr_out, avg_pump_pct), ...]  one tuple per step run
data = defaultdict(lambda: defaultdict(list))


def to_float(s):
    try:
        v = float(s)
        return None if math.isnan(v) else v
    except (ValueError, TypeError):
        return None


def flush_step(step_pct, step_stable):
    """Average all stable readings per hp_nom bucket for the completed step."""
    if step_pct is None:
        return
    for nom, readings in step_stable.items():
        if readings:
            avg_htr  = sum(r[0] for r in readings) / len(readings)
            avg_pump = sum(r[1] for r in readings) / len(readings)
            data[step_pct][nom].append((avg_htr, avg_pump))


def parse_sct(path):
    in_step     = False
    step_pct    = None
    skip_left   = 0
    stable_buf  = []               # rolling (h1, h2) for stability detection
    step_stable = defaultdict(list) # hp_nom → [(htr_out, pump)] during current step

    with open(path, 'r', errors='replace') as fh:
        for raw in fh:
            row = [c.strip() for c in raw.rstrip('\n').split(',')]
            if len(row) < 6:
                continue

            event = row[7].strip() if len(row) > 7 else ''

            if event == 'STEP_START':
                flush_step(step_pct, step_stable)
                pct = to_float(row[1])
                if pct is None:
                    continue
                step_pct    = int(round(pct))
                in_step     = True
                skip_left   = PRIME_ROWS
                stable_buf  = []
                step_stable = defaultdict(list)
                continue

            if event in ('STEP_END', 'PAUSE', 'OVERHEAT'):
                flush_step(step_pct, step_stable)
                step_stable = defaultdict(list)
                in_step = False
                continue

            if not in_step:
                continue

            if skip_left > 0:
                skip_left -= 1
                continue

            hp  = to_float(row[2])
            h1  = to_float(row[3])
            h2  = to_float(row[4])
            pmp = to_float(row[5])

            if any(v is None for v in (hp, h1, h2, pmp)):
                stable_buf = []
                continue

            stable_buf.append((h1, h2))
            if len(stable_buf) > STABLE_ROWS:
                stable_buf.pop(0)

            if len(stable_buf) < STABLE_ROWS:
                continue
            h1_range = max(x[0] for x in stable_buf) - min(x[0] for x in stable_buf)
            h2_range = max(x[1] for x in stable_buf) - min(x[1] for x in stable_buf)
            if h1_range >= STABLE_THR or h2_range >= STABLE_THR:
                continue

            nom     = int(round(hp))
            htr_out = max(h1, h2)
            if HP_MIN <= nom <= HP_MAX:
                step_stable[nom].append((htr_out, pmp))

    flush_step(step_pct, step_stable)  # commit any open step at EOF


for fn in sorted(os.listdir(LOG_DIR)):
    if fn.lower().startswith('sct') and fn.lower().endswith('.csv'):
        print(f"Parsing {fn} ...")
        parse_sct(os.path.join(LOG_DIR, fn))

EXCLUDE_LEVELS = {30, 70}  # old firmware levels superseded by 29% and 71%
HEATER_LEVELS = sorted(k for k in data.keys() if k not in EXCLUDE_LEVELS)

total_pts  = sum(len(v) for d in data.values() for v in d.values())
print(f"Parsed {total_pts} step averages across {len(HEATER_LEVELS)} heater levels.")
for pct in HEATER_LEVELS:
    if pct in data:
        d = data[pct]
        n = sum(len(v) for v in d.values())
        hp_min = min(d.keys()); hp_max = max(d.keys())
        print(f"  {pct:3d}%: {n:4d} steps, HP range {hp_min}-{hp_max}C")


# ── Pump_85 helpers ────────────────────────────────────────────────────────────

def pump85_linear(avg_htr, avg_pump, hp_nom):
    if hp_nom >= 85 or avg_htr <= hp_nom:
        return None
    result = avg_pump * (avg_htr - hp_nom) / (85.0 - hp_nom)
    return max(4.0, min(100.0, result))

def pump85_pl(avg_htr, avg_pump, hp_nom, pct):
    if hp_nom >= 85 or avg_htr <= hp_nom:
        return None
    alpha  = 1.534 - 0.734 * math.log(pct)
    ratio  = (85.0 - hp_nom) / (avg_htr - hp_nom)
    result = avg_pump * (ratio ** alpha)
    return max(4.0, min(100.0, result))


# ── Style helpers ──────────────────────────────────────────────────────────────

def solid_fill(hex_col):
    return PatternFill("solid", fgColor=hex_col)

def thin_border():
    s = Side(style="thin")
    return Border(left=s, right=s, top=s, bottom=s)

def hdr_cell(ws, row, col, text, bg="1F4E79", fg="FFFFFF"):
    c = ws.cell(row=row, column=col, value=text)
    c.font      = Font(bold=True, color=fg)
    c.fill      = solid_fill(bg)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border    = thin_border()
    return c

def data_cell(ws, row, col, value, bg=None, bold=False):
    c = ws.cell(row=row, column=col, value=value)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border    = thin_border()
    if bg:
        c.fill = solid_fill(bg)
    if bold:
        c.font = Font(bold=True)
    return c


# ── Build workbook ─────────────────────────────────────────────────────────────

wb = openpyxl.Workbook()
wb.remove(wb.active)

for pct in HEATER_LEVELS:
    sheet_data = data.get(pct, {})
    has_data   = bool(sheet_data)
    max_steps  = max((len(v) for v in sheet_data.values()), default=0)

    ws = wb.create_sheet(title=f"{pct}%")

    # ── Row 1: headers ─────────────────────────────────────────────────────────
    hdr_cell(ws, 1, 1, "HP (°C)")
    hdr_cell(ws, 1, 2, "Overall\nAvg")
    hdr_cell(ws, 1, 3, "Avg est\n@ 85°C", bg="375623")
    if has_data:
        c = ws.cell(row=1, column=4,
                    value=f"Each pair: observed avg  |  est. pump @ 85°C (midpoint)    steps={max_steps}")
        c.font      = Font(bold=True, color="FFFFFF")
        c.fill      = solid_fill("2E75B6")
        c.alignment = Alignment(horizontal="left", vertical="center")
        c.border    = thin_border()

    # ── Two rows per HP temp ────────────────────────────────────────────────────
    for hp_nom in range(HP_MIN, HP_MAX + 1):
        r_htr = (hp_nom - HP_MIN) * 2 + 2
        r_pmp = r_htr + 1

        steps = sheet_data.get(hp_nom, [])  # [(avg_htr, avg_pump), ...]

        # Overall averages
        overall_htr  = sum(s[0] for s in steps) / len(steps) if steps else None
        overall_pump = sum(s[1] for s in steps) / len(steps) if steps else None

        # Col C = average of per-step midpoint estimates
        midpoints = []
        for avg_htr, avg_pump in steps:
            lin = pump85_linear(avg_htr, avg_pump, hp_nom)
            if lin is not None:
                midpoints.append((avg_pump + lin) / 2.0)
        avg_est85 = sum(midpoints) / len(midpoints) if midpoints else None

        # ── htr_out row ────────────────────────────────────────────────────────
        data_cell(ws, r_htr, 1, "htr out", bg="FFF2CC", bold=True)
        data_cell(ws, r_htr, 2, round(overall_htr, 1) if overall_htr is not None else "", bg="FFF2CC")
        data_cell(ws, r_htr, 3, 85, bg="E2EFDA", bold=True)

        # Per-step pairs: col 4, 5 | 6, 7 | ...
        for i, (avg_htr, _) in enumerate(steps):
            col_obs = 4 + i * 2
            col_est = col_obs + 1
            data_cell(ws, r_htr, col_obs, round(avg_htr, 1), bg="FFF2CC")
            data_cell(ws, r_htr, col_est, 85, bg="E2EFDA", bold=True)

        # ── pump row ───────────────────────────────────────────────────────────
        data_cell(ws, r_pmp, 1, hp_nom, bg="D9E1F2", bold=True)
        data_cell(ws, r_pmp, 2, round(overall_pump, 1) if overall_pump is not None else "", bg="E2EFDA")
        if avg_est85 is not None:
            data_cell(ws, r_pmp, 3, round(avg_est85, 1), bg="C6EFCE", bold=True)
        else:
            data_cell(ws, r_pmp, 3, "", bg="D9E1F2")

        for i, (avg_htr, avg_pump) in enumerate(steps):
            col_obs = 4 + i * 2
            col_est = col_obs + 1
            lin = pump85_linear(avg_htr, avg_pump, hp_nom)
            data_cell(ws, r_pmp, col_obs, round(avg_pump, 1), bg="D9E1F2")
            if lin is not None:
                midpoint = (avg_pump + lin) / 2.0
                data_cell(ws, r_pmp, col_est, round(midpoint, 1), bg="C6EFCE")
            else:
                data_cell(ws, r_pmp, col_est, "", bg="D9E1F2")

    # ── Column widths ──────────────────────────────────────────────────────────
    ws.column_dimensions["A"].width = 9
    ws.column_dimensions["B"].width = 10
    ws.column_dimensions["C"].width = 10
    for i in range(max_steps):
        col_obs = 4 + i * 2
        col_est = col_obs + 1
        ws.column_dimensions[get_column_letter(col_obs)].width = 8
        ws.column_dimensions[get_column_letter(col_est)].width = 8

    ws.freeze_panes = "D2"
    ws.row_dimensions[1].height = 40

wb.save(OUT_PATH)
print(f"Saved: {OUT_PATH}")
